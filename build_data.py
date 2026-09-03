# 판매·재고·생산·수입·코드집 원본 5종을 읽어 PWA가 쓰는 data/data.json 을 만들고 검산 결과를 출력하는 빌드 스크립트
# -*- coding: utf-8 -*-
"""
사용법: update.bat 이 호출한다. 직접 실행해도 된다.
  python build_data.py

원본 경로는 전부 이 파일 안에서 관리한다 (배치파일에 한글·특수문자 경로를 넣으면 cmd가 깨진다).
컬럼 위치는 헤더 이름으로 찾는다 — 원본 열이 늘어나도 코드 수정 없이 동작해야 한다.
"""

import sys, os, re, io, json, glob, datetime, collections, urllib.request

# ── 경로 설정 ─────────────────────────────────────────────────────────
HERE = os.path.dirname(os.path.abspath(__file__))
if os.name == "nt":
    ROOT      = r"C:\Users\UNIX117\♥Claude"
    SALES     = os.path.join(ROOT, r"01)실적dashboard", "RAW_상품별유형별 실적_출력.xlsx")
    PROD_DIR  = os.path.join(ROOT, r"14)Notion", "01)제품생산계획 스페이스")
    STOCK_DIR = r"C:\★Jay\13.생산&재고\AA.재고_쿼리"
    CODEBOOK  = r"C:\★Jay\06.상품\▥상품코드집_운영.xlsx"
else:   # 개발용 샌드박스 마운트 경로
    M = "/sessions/peaceful-keen-cray/mnt"
    SALES     = f"{M}/01)실적dashboard/RAW_상품별유형별 실적_출력.xlsx"
    PROD_DIR  = f"{M}/01)제품생산계획 스페이스"
    STOCK_DIR = f"{M}/AA.재고_쿼리"
    CODEBOOK  = f"{M}/06.상품/▥상품코드집_운영.xlsx"

IMPORT_SHEET_ID  = "1P8lPE3Xx0RuwUCx1fuv19KYvrAy81f8eCRZwj46NHzI"      # 구글시트 '수입상품 입고일정'
IMPORT_URL       = f"https://docs.google.com/spreadsheets/d/{IMPORT_SHEET_ID}/export?format=xlsx"
IMPORT_TAB_PREFIX = "종합"                                              # 이 글자로 시작하는 탭을 쓴다

OUT_JSON  = os.path.join(HERE, "data", "data.json")
CACHE_DIR = os.path.join(HERE, "cache")

MONTHS_BACK   = 12     # 판매 표시 기간(최신 실적월 포함)
AVG_DAYS      = 90     # 회전일 계산용 일평균 판매 기간
IMPORT_PAST_D = 31     # 수입 일정: 오늘 기준 며칠 전까지 포함
SALES_SHEET   = "쿼리"
PROD_SHEET    = "일자별_raw"
CODE_SHEET    = "상품코드집"
# ─────────────────────────────────────────────────────────────────────

TODAY = datetime.date.today()


def ensure_openpyxl():
    try:
        import openpyxl  # noqa
    except ImportError:
        print("openpyxl 이 없어 자동 설치합니다...")
        os.system(f'"{sys.executable}" -m pip install openpyxl')
    import openpyxl
    return openpyxl


def ym_of(d):            # date → '2026.08'
    return f"{d.year}.{d.month:02d}"


def month_days(ym):      # '2026.08' → 그 달의 모든 날짜
    y, m = int(ym[:4]), int(ym[5:])
    n = (datetime.date(y + (m == 12), (m % 12) + 1, 1) - datetime.timedelta(days=1)).day
    return [datetime.date(y, m, i) for i in range(1, n + 1)]


def months_back(last_ym, n):
    y, m = int(last_ym[:4]), int(last_ym[5:])
    out = []
    for _ in range(n):
        out.append(f"{y}.{m:02d}")
        m -= 1
        if m == 0: y, m = y - 1, 12
    return out[::-1]


def norm(s):
    return str(s or "").strip()


# ── 1. 코드집 ────────────────────────────────────────────────────────
def read_codebook(openpyxl):
    """상품코드집 시트 → 제품명(B) 기준 순서·유형. A열 코드는 보조 키. 먼저 나온 행 우선."""
    if not os.path.exists(CODEBOOK):
        raise SystemExit(f"[중단] 코드집이 없습니다: {CODEBOOK}")
    ws = openpyxl.load_workbook(CODEBOOK, read_only=True, data_only=True)[CODE_SHEET]
    order, cat = {}, {}
    hdr = None
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if hdr is None:
            if r and "코드" in r and "제품명" in r and "대유형" in r:
                hdr = {n: r.index(n) for n in ("코드", "제품명", "대유형", "중유형", "소유형")}
            continue
        code, name = norm(r[hdr["코드"]]), norm(r[hdr["제품명"]])
        if not code and not name: continue
        c = (norm(r[hdr["대유형"]]), norm(r[hdr["중유형"]]), norm(r[hdr["소유형"]]))
        for k in (name, code):
            if k and k not in order:
                order[k] = i; cat[k] = c
    if hdr is None:
        raise SystemExit("[중단] 코드집 헤더(코드/제품명/대유형)를 찾지 못했습니다")
    return order, cat


# ── 2. 판매 ──────────────────────────────────────────────────────────
def read_sales(openpyxl):
    """쿼리 시트 → 완전중복 제거 후 (품명, 날짜, 분석채널, 거래처) 수량 집계."""
    if not os.path.exists(SALES):
        raise SystemExit(f"[중단] 판매 RAW가 없습니다: {SALES}")
    ws = openpyxl.load_workbook(SALES, read_only=True)[SALES_SHEET]
    it = ws.iter_rows(values_only=True)
    h = [norm(x) for x in next(it)]
    need = ["실적일자", "품명", "분석채널", "거래처", "수량", "대유형", "중유형", "소유형"]
    for n in need:
        if n not in h: raise SystemExit(f"[중단] 판매 RAW 헤더에 '{n}' 이 없습니다. 실제: {h}")
    ix = {n: h.index(n) for n in need}
    base = datetime.date(1899, 12, 30)
    seen, dup = set(), collections.Counter()
    agg = collections.Counter(); cat = {}; maxd = None; n = 0
    for r in it:
        n += 1
        key = tuple(str(x) for x in r)
        v = r[ix["실적일자"]]
        if isinstance(v, (int, float)):   d = base + datetime.timedelta(days=int(v))
        elif isinstance(v, datetime.datetime): d = v.date()
        elif isinstance(v, datetime.date): d = v
        else: continue
        if key in seen:
            dup[ym_of(d)] += 1; continue
        seen.add(key)
        pn = norm(r[ix["품명"]])
        if not pn: continue
        q = r[ix["수량"]] or 0
        agg[(pn, d, norm(r[ix["분석채널"]]), norm(r[ix["거래처"]]))] += q
        cat[pn] = (norm(r[ix["대유형"]]), norm(r[ix["중유형"]]), norm(r[ix["소유형"]]))
        maxd = d if (maxd is None or d > maxd) else maxd
    if dup:
        print(f"  [주의] 판매 RAW 완전중복 {sum(dup.values()):,}행 제거: " +
              ", ".join(f"{k} {v:,}" for k, v in sorted(dup.items())))
    print(f"  판매 RAW {n:,}행 · 최신 실적일 {maxd}")
    return agg, cat, maxd


# ── 3. 재고 ──────────────────────────────────────────────────────────
def read_stock(openpyxl, start):
    """start 이후 날짜의 창고별 재고조회 파일을 읽어 {날짜: {품명: 수량}}. 파일 단위 캐시(cache/stock_*.json)."""
    os.makedirs(CACHE_DIR, exist_ok=True)
    out, bad = {}, []
    files = sorted(glob.glob(os.path.join(STOCK_DIR, "창고별 재고조회_*.xlsx")))
    for path in files:
        m = re.search(r"_(\d{8})\.xlsx$", path)
        if not m or os.path.basename(path).startswith("~$"): continue
        d = datetime.date(int(m.group(1)[:4]), int(m.group(1)[4:6]), int(m.group(1)[6:]))
        if d < start: continue
        if os.path.getsize(path) == 0:
            bad.append((os.path.basename(path), "0바이트")); continue
        cp = os.path.join(CACHE_DIR, f"stock_{m.group(1)}.json")
        mt = os.path.getmtime(path)
        if os.path.exists(cp):
            try:
                c = json.load(open(cp, encoding="utf-8"))
                if abs(c.get("mtime", -1) - mt) < 1:
                    out[d] = c["inv"]; continue
            except Exception:
                pass
        try:
            ws = openpyxl.load_workbook(path, read_only=True, data_only=True).active
            rows = list(ws.iter_rows(values_only=True))
        except Exception as e:
            bad.append((os.path.basename(path), f"읽기 실패 {e}")); continue
        hi = None
        for i, r in enumerate(rows[:6]):
            if r and "품명" in r and "재고수량" in r: hi = i; break
        if hi is None:
            bad.append((os.path.basename(path), "헤더 없음")); continue
        cp_ = rows[hi].index("품명"); cq = rows[hi].index("재고수량")
        inv = collections.Counter(); total = None
        for r in rows[hi + 1:]:
            if not r: continue
            if norm(r[0]) == "TOTAL":
                total = r[cq]; continue
            pn = norm(r[cp_]); q = r[cq]
            if not pn or not isinstance(q, (int, float)): continue      # 2번째 헤더행 등 문자 셀 건너뜀
            inv[pn] += q
        if total is not None and sum(inv.values()) != total:
            bad.append((os.path.basename(path), f"합계 불일치 {sum(inv.values())} vs TOTAL {total}"))
        out[d] = dict(inv)
        json.dump({"mtime": mt, "inv": out[d]}, open(cp, "w", encoding="utf-8"), ensure_ascii=False)
    if bad:
        print("  [주의] 재고 파일 문제:", "; ".join(f"{a} ({b})" for a, b in bad))
    print(f"  재고 파일 {len(out)}일치 ({min(out) if out else '-'} ~ {max(out) if out else '-'})")
    return out


# ── 4. 생산계획 ──────────────────────────────────────────────────────
def read_prod(openpyxl):
    """가장 최신 일자별_자재계획_*.xlsx 의 일자별_raw → 오늘 이후 계획수량>0 행."""
    cands = [p for p in glob.glob(os.path.join(PROD_DIR, "일자별_자재계획_*.xlsx"))
             if not os.path.basename(p).startswith("~$")]
    if not cands:
        print("  [주의] 생산계획 파일이 없습니다 — 생산 일정 없이 진행"); return [], None
    path = sorted(cands, key=os.path.getmtime)[-1]
    ws = openpyxl.load_workbook(path, read_only=True, data_only=True)[PROD_SHEET]
    it = ws.iter_rows(values_only=True)
    h = [norm(x) for x in next(it)]
    for n in ("생산일자", "모델", "계획수량"):
        if n not in h: raise SystemExit(f"[중단] 생산계획 헤더에 '{n}' 없음: {h}")
    ix = {n: h.index(n) for n in ("생산일자", "모델", "계획수량")}
    rows = []; total = 0
    for r in it:
        d = r[ix["생산일자"]]
        if isinstance(d, datetime.datetime): d = d.date()
        if not isinstance(d, datetime.date): continue
        q = int(r[ix["계획수량"]] or 0); total += q
        if d >= TODAY and q > 0:
            rows.append((norm(r[ix["모델"]]), d, q))
    print(f"  생산계획 {os.path.basename(path)} · 오늘 이후 {len(rows)}건 (파일 전체 계획합 {total:,})")
    return rows, os.path.basename(path)


# ── 5. 수입 일정 ─────────────────────────────────────────────────────
def download_import():
    """구글시트를 xlsx로 내려받아 cache/에 저장. 실패하면 마지막 성공본 경로를 돌려준다."""
    os.makedirs(CACHE_DIR, exist_ok=True)
    latest = os.path.join(CACHE_DIR, "import_latest.xlsx")
    try:
        req = urllib.request.Request(IMPORT_URL, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=60) as resp:
            data = resp.read()
        if not data.startswith(b"PK"):
            raise RuntimeError("xlsx 형식이 아닌 응답 (공유 설정이 바뀌었을 수 있음)")
        open(latest, "wb").write(data)
        open(os.path.join(CACHE_DIR, f"import_{TODAY:%Y%m%d}.xlsx"), "wb").write(data)
        return latest, f"{datetime.datetime.now():%Y-%m-%d %H:%M} 다운로드"
    except Exception as e:
        if os.path.exists(latest):
            t = datetime.datetime.fromtimestamp(os.path.getmtime(latest))
            print(f"  [주의] 구글시트 다운로드 실패 ({e}) → 마지막 성공본 사용 ({t:%Y-%m-%d %H:%M})")
            return latest, f"캐시 {t:%Y-%m-%d %H:%M}"
        print(f"  [주의] 구글시트 다운로드 실패 ({e}) · 캐시도 없음 → 수입 일정 없이 진행")
        return None, "없음"


_MD = re.compile(r"(\d{1,2})\s*[월/.]\s*(\d{1,2})")
_YMD = re.compile(r"(20\d{2})\D{1,3}(\d{1,2})\D{1,3}(\d{1,2})")


def parse_date(v, year, month_hint):
    """'1월11일' · '1/25' · '2024. 1. 13 오전' · datetime → date. 연도는 월 헤더 기준, 12월→1월 넘김 보정."""
    if v is None: return None
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, datetime.date): return v
    s = str(v)
    m = _YMD.search(s)
    if m:
        try: return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError: return None
    m = _MD.search(s)
    if not m: return None
    mo, da = int(m.group(1)), int(m.group(2))
    if not (1 <= mo <= 12 and 1 <= da <= 31): return None
    y = year
    if month_hint and mo < month_hint - 6: y += 1          # 12월 헤더 아래 1월 ETA
    try: return datetime.date(y, mo, da)
    except ValueError: return None


def parse_qty(v):
    if v is None: return 0
    if isinstance(v, (int, float)): return int(v)
    s = re.sub(r"[^\d\-]", "", str(v))
    return int(s) if s.lstrip("-").isdigit() else 0


def read_import(openpyxl, path):
    """종합 탭: 월 헤더('N월,YYYY년') 아래 NO. 행 + 이어지는 MODEL 행. 일자 = 김포입고일 우선, 없으면 ETA."""
    if not path: return [], collections.Counter()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    name = next((s for s in wb.sheetnames if s.startswith(IMPORT_TAB_PREFIX)), None)
    if not name:
        print(f"  [주의] '{IMPORT_TAB_PREFIX}' 탭이 없습니다: {wb.sheetnames}"); return [], collections.Counter()
    rows = list(wb[name].iter_rows(values_only=True))
    hi = next((i for i, r in enumerate(rows) if r and any(norm(x) == "MODEL" for x in r)), None)
    if hi is None:
        print("  [주의] 수입 탭에서 MODEL 헤더를 못 찾았습니다"); return [], collections.Counter()
    h1 = [norm(x) for x in rows[hi]]; h2 = [norm(x) for x in rows[hi + 1]] if hi + 1 < len(rows) else []
    c_no, c_model, c_qty = h1.index("NO."), h1.index("MODEL"), h1.index("QTY")
    c_eta = h2.index("ETA") if "ETA" in h2 else None
    c_gp = next((i for i, x in enumerate(h1) if "김포" in x), None)
    year = month = None; cur = None
    out = []; stat = collections.Counter(); cutoff = TODAY - datetime.timedelta(days=IMPORT_PAST_D)
    for r in rows[hi + 2:]:
        if not r: continue
        a = norm(r[0])
        mh = re.fullmatch(r"(\d{1,2})월\s*,\s*(\d{4})년", a)
        if mh:
            month, year = int(mh.group(1)), int(mh.group(2)); cur = None; continue
        if year is None: continue
        if a and c_no is not None:
            no = a
            eta = parse_date(r[c_eta], year, month) if c_eta is not None else None
            gp = parse_date(r[c_gp], year, month) if c_gp is not None else None
            cur = (no, eta, gp)
        if cur is None: continue
        model = norm(r[c_model]) if c_model < len(r) else ""
        if not model: continue
        stat["행"] += 1
        no, eta, gp = cur
        d, basis = (gp, "김포") if gp else ((eta, "ETA") if eta else (None, None))
        if d is None: stat["일자없음"] += 1; continue
        if d < cutoff: stat["기간외"] += 1; continue
        out.append((model, d, parse_qty(r[c_qty]), basis, no))
    print(f"  수입 일정 '{name}' 탭 · MODEL 행 {stat['행']:,} · 기간 내 {len(out)} · 일자 없음 {stat['일자없음']} · 기간 외 {stat['기간외']}")
    return out, stat


# ── 6. 조립 ──────────────────────────────────────────────────────────
def build():
    openpyxl = ensure_openpyxl()
    print("=" * 64); print(f"stock-supply 데이터 빌드  {datetime.datetime.now():%Y-%m-%d %H:%M}"); print("=" * 64)

    print("[1] 코드집");  order, cb_cat = read_codebook(openpyxl)
    print(f"  코드집 키 {len(order):,}개")
    print("[2] 판매");    sales, s_cat, maxd = read_sales(openpyxl)
    months = months_back(ym_of(maxd), MONTHS_BACK)
    start = month_days(months[0])[0]
    print(f"  표시 기간 {months[0]} ~ {months[-1]}")
    print("[3] 재고");    stock = read_stock(openpyxl, start - datetime.timedelta(days=7))
    print("[4] 생산계획"); prod, prod_name = read_prod(openpyxl)
    print("[5] 수입 일정")
    imp_path, imp_src = download_import()
    imp, imp_stat = read_import(openpyxl, imp_path)

    # 품목 우주 = 코드집 ∪ 판매실적(표시 기간)
    win_sales = {k: v for k, v in sales.items() if k[1] >= start}
    sold = {k[0] for k in win_sales}
    universe = set(order) | sold
    stock_days = sorted(stock)
    latest_day = stock_days[-1] if stock_days else None
    cur = stock.get(latest_day, {}) if latest_day else {}

    # 수급: 품명별
    sup = collections.defaultdict(list)
    for m, d, q in prod: sup[m].append(["생산", d.isoformat(), q, ""])
    imp_unmatched = collections.Counter()
    for m, d, q, basis, no in imp:
        if m in universe: sup[m].append(["수입", d.isoformat(), q, basis, no])
        else: imp_unmatched[m] += 1
    for v in sup.values(): v.sort(key=lambda x: x[1])

    # 90일 평균
    a_start = maxd - datetime.timedelta(days=AVG_DAYS - 1)
    avg = collections.Counter()
    for (pn, d, ch, cu), q in sales.items():
        if d >= a_start: avg[pn] += q

    # 사전
    ch_list = sorted({k[2] for k in win_sales}); cu_list = sorted({k[3] for k in win_sales})
    ch_ix = {v: i for i, v in enumerate(ch_list)}; cu_ix = {v: i for i, v in enumerate(cu_list)}

    days = {ym: [d.isoformat() for d in month_days(ym)] for ym in months}
    # 재고 일자 배열은 파일이 있는 날만 값, 없으면 null. 전월 말일(직전 파일일) 재고도 함께
    prev_stock = {}
    for ym in months:
        first = month_days(ym)[0]
        before = [d for d in stock_days if d < first]
        prev_stock[ym] = before[-1] if before else None

    by_pn = collections.defaultdict(list)
    for (p, d, ch, cu), q in win_sales.items(): by_pn[p].append((d, ch, cu, q))

    items = []; dropped_no_activity = 0
    for pn in sorted(universe):
        det = {}; sm = {}; has_sale = False
        for ym in months:
            sm[ym] = [0] * len(days[ym])
        for d, ch, cu, q in by_pn.get(pn, ()):
            ym = ym_of(d); i = d.day - 1
            sm[ym][i] += q; has_sale = True
            det.setdefault(ym, {}).setdefault(str(d.day), []).append([ch_ix[ch], cu_ix[cu], q])
        for ym in det:
            for k in det[ym]: det[ym][k].sort(key=lambda x: -x[2])
        stk = {}; k0 = {}
        for ym in months:
            arr = []
            for d in month_days(ym):
                if d in stock: arr.append(stock[d].get(pn, 0))
                else: arr.append(None)
            if any(v is not None for v in arr): stk[ym] = arr
            pd = prev_stock[ym]
            if pd: k0[ym] = stock[pd].get(pn, 0)
        c = cur.get(pn, 0)
        if not has_sale and c == 0 and pn not in sup:
            dropped_no_activity += 1; continue          # 판매·재고·수급 모두 없는 품목 숨김
        cat = s_cat.get(pn) or cb_cat.get(pn) or ("", "", "")
        items.append({
            "pn": pn, "cat": list(cat), "ord": order.get(pn, 10 ** 6),
            "cur": c, "avg": round(avg[pn] / AVG_DAYS, 2),
            "s": {ym: sm[ym] for ym in months if any(sm[ym])},
            "d": det, "k": stk, "k0": k0, "sup": sup.get(pn, []),
        })
    items.sort(key=lambda x: (x["ord"], x["pn"]))

    data = {
        "built": f"{datetime.datetime.now():%Y-%m-%d %H:%M}",
        "asof": {"sales": maxd.isoformat(), "stock": latest_day.isoformat() if latest_day else None,
                 "prod": prod_name or "-", "imp": imp_src, "today": TODAY.isoformat()},
        "months": months, "days": days, "ch": ch_list, "cu": cu_list,
        "rules": {"avgDays": AVG_DAYS, "inflow": 30, "lowTurn": 15},
        "items": items,
    }

    # ── 검산 ──
    print("[6] 검산")
    ok = True
    tot_json = sum(sum(v) for it in items for v in it["s"].values())
    tot_raw = sum(q for (p, d, ch, cu), q in win_sales.items() if p in universe)
    tot_raw_all = sum(win_sales.values())
    print(f"  판매 합계  JSON {tot_json:,} = RAW(우주 내) {tot_raw:,} : {'OK' if tot_json == tot_raw else 'NG'}"
          + (f"  (우주 밖 품명 판매 {tot_raw_all - tot_raw:,} 제외)" if tot_raw_all != tot_raw else ""))
    ok &= tot_json == tot_raw
    if latest_day:
        js = sum(it["cur"] for it in items); rw = sum(cur.values())
        shown = {it["pn"] for it in items}
        excl = sum(v for k, v in cur.items() if k not in shown)
        print(f"  재고 최신({latest_day}) JSON {js:,} + 제외품목 {excl:,} = 파일 {rw:,} : {'OK' if js + excl == rw else 'NG'}")
        ok &= js + excl == rw
    pj = sum(x[2] for it in items for x in it["sup"] if x[0] == "생산"); pr = sum(q for m, d, q in prod if m in universe)
    print(f"  생산(오늘 이후) JSON {pj:,} = 파일 {pr:,} : {'OK' if pj == pr else 'NG'}  (우주 밖 모델 {sum(q for m,d,q in prod if m not in universe):,})")
    ok &= pj == pr
    ij = sum(1 for it in items for x in it["sup"] if x[0] == "수입")
    print(f"  수입 매칭 {ij}건 · 미매칭 MODEL {len(imp_unmatched)}종"
          + (": " + ", ".join(list(imp_unmatched)[:12]) if imp_unmatched else ""))
    for ym in months:
        a = sum(sum(it["s"].get(ym, [])) for it in items)
        b = sum(q for (p, d, ch, cu), q in win_sales.items() if ym_of(d) == ym and p in universe)
        if a != b: print(f"  [NG] {ym} 판매 {a:,} vs {b:,}"); ok = False
    print(f"  품목 {len(items):,}개 표시 · 활동 없음 숨김 {dropped_no_activity:,} · 코드집 미등록(뒤에 정렬) {sum(1 for it in items if it['ord']>=10**6)}")
    if not ok:
        raise SystemExit("[중단] 검산 불일치 — data.json 을 만들지 않았습니다")

    os.makedirs(os.path.dirname(OUT_JSON), exist_ok=True)
    js = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    io.open(OUT_JSON, "w", encoding="utf-8").write(js)
    print(f"\nOK  data/data.json {len(js.encode('utf-8'))/1024:,.0f} KB · 기준 판매 {maxd} · 재고 {latest_day} · 생산 {prod_name} · 수입 {imp_src}")


if __name__ == "__main__":
    build()
